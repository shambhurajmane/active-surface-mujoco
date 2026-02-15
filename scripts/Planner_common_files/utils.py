import os
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import time


def get_config_from_yaml(yaml_file_path=r"configs\config.yaml"):
    # yaml_file_path = os.path.join(rospack.get_path('vf_hand'), 'config', 'config.yaml')
    with open(yaml_file_path, 'r') as file:
        try:
            configs = yaml.safe_load(file)
            return configs
        except yaml.YAMLError as exc:
            print(f"Error reading YAML file: {exc}")
            return None
        
        


def save_data(object_name, id, action_sequence, action_list ,no_of_switches, pivot_count, rotation_count, time_taken, states_explored):
    file_location =   r"C:\Users\shamb\sam\PHD\WPI\PHD\ScienceRobotics\MujoCo\Release\Results\planning_results.xlsx"  
    # check if file exists
    # if not create a new file and write the data
    sheet = "Sheet1"
    
    row = None  
    if not os.path.exists(file_location) or row !=None:
        # Create a new workbook and add a sheet with headers
        workbook = Workbook()
        # create a new sheet
        worksheet = workbook.active
        worksheet.title = sheet
        row = 1
        col =1

        worksheet.cell(row, col, "object_name")
        col = 2
        worksheet.cell(row, col, "id")
        worksheet.cell(row, col, "Heuristic_used")
        worksheet.cell(row, col+1, "action_list")
        worksheet.cell(row, col+1, "no_of_switches")
        worksheet.cell(row, col+1, "pivot_count")
        worksheet.cell(row, col+1, "rotation_count")
        worksheet.cell(row, col+1, "time_taken")
        worksheet.cell(row, col+1, "states_explored")  
        # cost function is # Cost of each action to account for in A* algorithm
        row += 1
    else:
        # Open the existing workbook and get the active sheet
        workbook = load_workbook(filename=file_location)
        worksheet = workbook[sheet]
        # Get the number of rows and set the row variable to append data
        data = worksheet.iter_rows(values_only=True)
        row = 1
        col =1
        for row_data in data:
            worksheet.cell(row, col, row_data[0])   
            worksheet.cell(row, col+1, row_data[1]) 
            worksheet.cell(row, col+2, row_data[2])
            worksheet.cell(row, col+3, row_data[3])
            worksheet.cell(row, col+4, row_data[4])
            worksheet.cell(row, col+5, row_data[5])
            worksheet.cell(row, col+6, row_data[6])
            worksheet.cell(row, col+7, row_data[7])    
            row += 1
        # object_name, id, action_list ,no_of_switches, pivot_count, rotation_count, time_taken, states_explored
        col = 1
        worksheet.cell(row, col, object_name)
        col = 2
        worksheet.cell(row, col, id)  
        worksheet.cell(row, col+1, "l1_norm")
        worksheet.cell(row, col+2, str(action_list))
        worksheet.cell(row, col+3, no_of_switches)
        worksheet.cell(row, col+4, pivot_count)
        worksheet.cell(row, col+5, rotation_count)
        worksheet.cell(row, col+6, time_taken)
        worksheet.cell(row, col+7, states_explored)# Save the workbook

        workbook.save(file_location)
        print(f"Data saved to {file_location}")